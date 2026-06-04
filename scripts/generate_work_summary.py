#!/usr/bin/env python3

from __future__ import annotations

import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = REPO_ROOT / "reports" / "may-june-work-summary.xlsx"
DATE_START = "2026-05-01"
DATE_END = "2026-06-30 23:59:59"


@dataclass
class CommitRecord:
    commit: str
    date: str
    author: str
    subject: str
    files_changed: int
    insertions: int
    deletions: int
    area: str
    category: str
    before_state: str
    after_state: str
    improvement: str


def run_git_log() -> str:
    cmd = [
        "git",
        "log",
        "--no-merges",
        f"--since={DATE_START}",
        f"--until={DATE_END}",
        "--date=short",
        "--numstat",
        "--pretty=format:===COMMIT===%n%H%n%ad%n%an%n%s",
    ]
    result = subprocess.run(
        cmd,
        cwd=REPO_ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


def normalize_subject(subject: str) -> str:
    return re.sub(r"^(feat|fix|docs|chore|test|ci):\s*", "", subject, flags=re.IGNORECASE)


def categorize(subject: str) -> str:
    lowered = subject.lower()
    if lowered.startswith(("test:", "docs:", "chore:", "ci:")):
        return "Other"
    if lowered.startswith("feat:"):
        return "Feature"
    if lowered.startswith("fix:") or lowered.startswith("fix ") or lowered.startswith("harden "):
        return "Fix"
    if lowered.startswith("add "):
        if any(token in lowered for token in ["test", "tests", "docs", "documentation", "skill", "workflow"]):
            return "Other"
        return "Feature"
    return "Other"


def infer_area(files: list[str], subject: str) -> str:
    lowered = subject.lower()
    joined = " ".join(files).lower()
    if ".agents/skills/" in joined or "skill" in lowered:
        return "Agent Tooling"
    if "analytics" in lowered or "analytics" in joined or "posthog" in lowered:
        return "Analytics"
    if "export" in lowered or "/exports" in joined:
        return "Exports"
    if "notification" in lowered or "notification" in joined:
        return "Notifications"
    if "prisma" in lowered or "prisma" in joined:
        return "Database / Prisma"
    if "sql" in lowered or ".sql" in joined:
        return "Database / SQL"
    if "test" in lowered or "tests/" in joined or "vitest" in lowered:
        return "Testing / CI"
    if "doc" in lowered or "readme" in joined or "documentation/" in joined:
        return "Documentation"
    if "filter" in lowered or "search" in lowered:
        return "Search / Filters"
    if "favorite" in lowered or "selection" in lowered:
        return "Dashboard UX"
    if "etl" in lowered or "etl/" in joined or "main_with_notifications.py" in joined:
        return "ETL / Data Pipeline"
    if "dashboard" in lowered or "summary" in lowered:
        return "Dashboard"
    return "General"


def sentence_case(text: str) -> str:
    text = text.strip()
    if not text:
        return text
    return text[0].upper() + text[1:]


def build_before_after(subject: str, category: str, area: str) -> tuple[str, str, str]:
    action = normalize_subject(subject)
    before = f"Prior state in {area.lower()} did not include this change or was behaving inconsistently."
    after = f"Implemented: {sentence_case(action)}."
    if category == "Feature":
        improvement = "Adds new user-facing capability and broadens supported workflows."
    elif category == "Fix":
        improvement = "Reduces incorrect behavior, edge-case failures, or production risk."
    else:
        improvement = "Improves maintainability, observability, test confidence, documentation, or platform readiness."

    lowered = subject.lower()
    if "coverage" in lowered or "test" in lowered:
        before = "Automated regression coverage was incomplete, leaving more behavior unverified."
        after = f"Expanded test coverage for: {action}."
        improvement = "Raises confidence in releases and catches regressions earlier."
    elif "export" in lowered:
        before = "Export workflows had weaker controls, broader scope, or less predictable behavior."
        after = f"Updated export flow to: {action}."
        improvement = "Improves export reliability, correctness, and operational safety."
    elif "analytics" in lowered or "posthog" in lowered:
        before = "Product interaction tracking captured less context or had review gaps."
        after = f"Adjusted analytics implementation to: {action}."
        improvement = "Improves product visibility and makes usage analysis more actionable."
    elif "prisma" in lowered:
        before = "Database access relied on the previous integration path with less standardization."
        after = f"Shifted database layer to: {action}."
        improvement = "Improves maintainability, data-layer consistency, and future schema evolution."
    elif "sql" in lowered or "query" in lowered or "load" in lowered:
        before = "Relevant queries were doing more work than necessary or scaling less efficiently."
        after = f"Optimized data path to: {action}."
        improvement = "Improves response time and lowers backend processing cost."
    elif "docs" in lowered or "documentation" in lowered:
        before = "Project guidance or architecture references were incomplete or outdated."
        after = f"Updated documentation to: {action}."
        improvement = "Improves onboarding speed and reduces implementation ambiguity."
    elif "fix" in lowered or "harden" in lowered:
        before = "An incorrect, fragile, or risky behavior existed in the current flow."
        after = f"Corrected behavior to: {action}."
        improvement = "Improves correctness and reduces production issues."

    return before, after, improvement


def parse_git_log(raw: str) -> list[CommitRecord]:
    records: list[CommitRecord] = []
    chunks = [chunk.strip() for chunk in raw.split("===COMMIT===") if chunk.strip()]
    for chunk in chunks:
        lines = chunk.splitlines()
        if len(lines) < 4:
            continue
        commit, date, author, subject = lines[:4]
        files: list[str] = []
        insertions = 0
        deletions = 0
        for line in lines[4:]:
            stripped = line.strip()
            if not stripped:
                continue
            parts = stripped.split("\t")
            if len(parts) != 3:
                continue
            added, removed, path = parts
            files.append(path)
            if added.isdigit():
                insertions += int(added)
            if removed.isdigit():
                deletions += int(removed)

        area = infer_area(files, subject)
        category = categorize(subject)
        before_state, after_state, improvement = build_before_after(subject, category, area)
        records.append(
            CommitRecord(
                commit=commit,
                date=date,
                author=author,
                subject=subject,
                files_changed=len(files),
                insertions=insertions,
                deletions=deletions,
                area=area,
                category=category,
                before_state=before_state,
                after_state=after_state,
                improvement=improvement,
            )
        )
    return sorted(records, key=lambda record: (record.date, record.category, record.subject))


def autosize(ws) -> None:
    widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths[cell.column], len(value) + 2), 60)
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    autosize(ws)


def add_summary_sheet(wb: Workbook, records: list[CommitRecord]) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(
        [
            "Month",
            "Days With Work",
            "Entries",
            "Feature",
            "Fix",
            "Other",
            "Top Areas",
        ]
    )

    by_month: dict[str, list[CommitRecord]] = defaultdict(list)
    for record in records:
        month = datetime.strptime(record.date, "%Y-%m-%d").strftime("%B %Y")
        by_month[month].append(record)

    for month in ["May 2026", "June 2026"]:
        month_records = by_month.get(month, [])
        category_counts = Counter(record.category for record in month_records)
        area_counts = Counter(record.area for record in month_records)
        top_areas = ", ".join(area for area, _ in area_counts.most_common(3))
        ws.append(
            [
                month,
                len({record.date for record in month_records}),
                len(month_records),
                category_counts.get("Feature", 0),
                category_counts.get("Fix", 0),
                category_counts.get("Other", 0),
                top_areas,
            ]
        )

    style_sheet(ws)


def add_month_sheet(wb: Workbook, month: int, name: str, records: list[CommitRecord]) -> None:
    ws = wb.create_sheet(title=name)
    ws.append(
        [
            "Date",
            "Category",
            "Area",
            "Commit Summary",
            "Before",
            "After",
            "Improvement",
            "Files Changed",
            "Insertions",
            "Deletions",
            "Author",
            "Commit",
        ]
    )

    for record in records:
        record_month = datetime.strptime(record.date, "%Y-%m-%d").month
        if record_month != month:
            continue
        ws.append(
            [
                record.date,
                record.category,
                record.area,
                sentence_case(normalize_subject(record.subject)),
                record.before_state,
                record.after_state,
                record.improvement,
                record.files_changed,
                record.insertions,
                record.deletions,
                record.author,
                record.commit[:10],
            ]
        )

    style_sheet(ws)
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    raw = run_git_log()
    records = parse_git_log(raw)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    add_summary_sheet(wb, records)
    add_month_sheet(wb, 5, "May 2026", records)
    add_month_sheet(wb, 6, "June 2026", records)
    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
